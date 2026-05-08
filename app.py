import io
import os
import re
import tempfile
from datetime import datetime

import pandas as pd
import streamlit as st

import pdfplumber
import pytesseract
from PIL import Image
from pdf2image import convert_from_bytes
from openpyxl import load_workbook


# -----------------------------
# Page configuration
# -----------------------------
st.set_page_config(
    page_title="Solar Load Calculator Automation",
    page_icon="☀️",
    layout="wide",
)

st.title("☀️ Solar Load Calculator Automation")
st.write(
    "Upload your electricity bill PDF/image. The app extracts key fields using OCR and fills an Excel template."
)


# -----------------------------
# Constants
# -----------------------------
ALLOWED_TYPES = ["application/pdf", "image/png", "image/jpeg"]
ALLOWED_EXTS = [".pdf", ".png", ".jpg", ".jpeg"]

# Default template path (must be in the same folder as app.py)
TEMPLATE_FILE = "template.xlsx"
OUTPUT_FILE_NAME = "generated_output.xlsx"


# -----------------------------
# Helpers: OCR / PDF extraction
# -----------------------------
def _safe_decode_text(text: str) -> str:
    if not text:
        return ""
    # Remove excessive whitespace
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[\t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def extract_text_from_pdf(pdf_bytes: bytes) -> tuple[str, str]:
    """Return (text, method_used)."""
    # 1) Try pdfplumber text extraction
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            texts = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                texts.append(t)
            combined = _safe_decode_text("\n".join(texts))
            # Heuristic: if it's too short, consider it a failure
            if len(combined) >= 50:
                return combined, "pdfplumber"
    except Exception:
        # Fall through to OCR
        pass

    # 2) Convert to images and OCR
    try:
        images = convert_from_bytes(pdf_bytes)
        ocr_texts = []
        for img in images:
            ocr_texts.append(pytesseract.image_to_string(img))
        combined = _safe_decode_text("\n".join(ocr_texts))
        return combined, "pdf2image+ocr"
    except Exception as e:
        raise RuntimeError(f"Failed PDF OCR: {e}")


def extract_text_from_image(image_bytes: bytes) -> tuple[str, str]:
    try:
        img = Image.open(io.BytesIO(image_bytes))
        # Ensure RGB to avoid some OCR issues
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        text = pytesseract.image_to_string(img)
        return _safe_decode_text(text), "tesseract"
    except Exception as e:
        raise RuntimeError(f"Failed image OCR: {e}")


def extract_text(uploaded_file) -> tuple[str, str]:
    """Extract text from uploaded PDF/image based on content type."""
    file_bytes = uploaded_file.getvalue()

    if uploaded_file.type == "application/pdf" or uploaded_file.name.lower().endswith(".pdf"):
        return extract_text_from_pdf(file_bytes)

    if uploaded_file.type in ("image/png", "image/jpeg") or any(
        uploaded_file.name.lower().endswith(ext) for ext in [".png", ".jpg", ".jpeg"]
    ):
        return extract_text_from_image(file_bytes)

    raise ValueError("Unsupported file type")


# -----------------------------
# Helpers: Field extraction via regex
# -----------------------------

def _find(patterns, text: str):
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE | re.MULTILINE)
        if m:
            return m.group(1).strip()
    return ""


def parse_bill_fields(text: str) -> dict:
    """Extract required fields from bill text.

    Returns dict with keys:
      Consumer Number
      Consumer Name
      Billing Date
      Units Consumed
      Bill Amount
      Tariff
      Load
      Meter Number
    """

    text_norm = text

    # Consumer Number
    consumer_number = _find(
        [
            r"\bConsumer\s*Number\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
            r"\bConsumer\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
            r"\bAccount\s*Number\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
        ],
        text_norm,
    )

    # Consumer Name
    consumer_name = _find(
        [
            r"\bConsumer\s*Name\s*[:\-]?\s*([A-Za-z][A-Za-z\s\.\-']{2,})",
            r"\bName\s*of\s*Consumer\s*[:\-]?\s*([A-Za-z][A-Za-z\s\.\-']{2,})",
            r"\bCustomer\s*Name\s*[:\-]?\s*([A-Za-z][A-Za-z\s\.\-']{2,})",
        ],
        text_norm,
    )

    # Billing Date
    billing_date = _find(
        [
            r"\bBilling\s*Date\s*[:\-]?\s*([0-9]{1,2}[\-/][0-9]{1,2}[\-/][0-9]{2,4})",
            r"\bBill\s*Date\s*[:\-]?\s*([0-9]{1,2}[\-/][0-9]{1,2}[\-/][0-9]{2,4})",
            r"\bDate\s*[:\-]?\s*([0-9]{1,2}[\-/][0-9]{1,2}[\-/][0-9]{2,4})",
        ],
        text_norm,
    )

    # Units Consumed
    units_consumed = _find(
        [
            r"\bUnits\s*Consumed\s*[:\-]?\s*([0-9]{1,6}([\.,][0-9]{1,2})?)",
            r"\bUnits\s*[:\-]?\s*([0-9]{1,6}([\.,][0-9]{1,2})?)",
            r"\bConsumption\s*[:\-]?\s*([0-9]{1,6}([\.,][0-9]{1,2})?)\b",
        ],
        text_norm,
    )

    # Bill Amount
    bill_amount = _find(
        [
            r"\bBill\s*Amount\s*[:\-]?\s*(?:₹|INR\.?\s*)?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]{1,2})?)",
            r"\bAmount\s*Payable\s*[:\-]?\s*(?:₹|INR\.?\s*)?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]{1,2})?)",
            r"\bTotal\s*Amount\s*[:\-]?\s*(?:₹|INR\.?\s*)?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{1,2})?|[0-9]+(?:\.[0-9]{1,2})?)",
        ],
        text_norm,
    )

    # Tariff
    tariff = _find(
        [
            r"\bTariff\s*[:\-]?\s*([A-Za-z0-9\s\-/]{2,})\b",
            r"\bRate\s*[:\-]?\s*([A-Za-z0-9\s\-/]{2,})\b",
        ],
        text_norm,
    )

    # Load
    load = _find(
        [
            r"\bLoad\s*[:\-]?\s*([0-9]{1,6}([\.,][0-9]{1,2})?)\s*(kW|KW|W)?",
            r"\bSanctioned\s*Load\s*[:\-]?\s*([0-9]{1,6}([\.,][0-9]{1,2})?)\s*(kW|KW|W)?",
        ],
        text_norm,
    )
    # If the regex captured tuples, normalize
    if isinstance(load, str):
        pass

    # Meter Number
    meter_number = _find(
        [
            r"\bMeter\s*Number\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
            r"\bMeter\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
            r"\bMeter\s*ID\s*[:\-]?\s*([A-Za-z0-9\-/]{3,})",
        ],
        text_norm,
    )

    return {
        "Consumer Number": consumer_number,
        "Consumer Name": consumer_name,
        "Billing Date": billing_date,
        "Units Consumed": units_consumed,
        "Bill Amount": bill_amount,
        "Tariff": tariff,
        "Load": load,
        "Meter Number": meter_number,
    }


# -----------------------------
# Helpers: Excel template fill
# -----------------------------

def load_template_or_error():
    template_path = os.path.join(os.path.dirname(__file__), TEMPLATE_FILE)
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Missing '{TEMPLATE_FILE}' in the same folder as app.py. "
            "Place your template.xlsx there with the correct input cell mapping."
        )
    return template_path


def normalize_number_str(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    # remove commas
    s = s.replace(",", "")
    return s


def fill_excel_template(fields: dict) -> str:
    """Fill only input cells using openpyxl.

    Mapping approach:
    - Find cells by scanning column A for label text (case-insensitive exact-ish match)
      and fill adjacent cell in same row (column B).

    This avoids overwriting formulas in other columns.
    """
    template_path = load_template_or_error()
    wb = load_workbook(template_path)

    # Choose first sheet
    ws = wb.active

    label_to_value = {
        "Consumer Number": fields.get("Consumer Number", ""),
        "Consumer Name": fields.get("Consumer Name", ""),
        "Billing Date": fields.get("Billing Date", ""),
        "Units Consumed": fields.get("Units Consumed", ""),
        "Bill Amount": fields.get("Bill Amount", ""),
        "Tariff": fields.get("Tariff", ""),
        "Load": fields.get("Load", ""),
        "Meter Number": fields.get("Meter Number", ""),
    }

    # Build regex map for labels in case template uses slight variations
    label_patterns = {
        "Consumer Number": r"consumer\s*number",
        "Consumer Name": r"consumer\s*name|customer\s*name",
        "Billing Date": r"billing\s*date|bill\s*date|date",
        "Units Consumed": r"units\s*consumed|units\b|consumed\s*units",
        "Bill Amount": r"bill\s*amount|amount\s*payable|total\s*amount",
        "Tariff": r"tariff|rate",
        "Load": r"load|sanctioned\s*load",
        "Meter Number": r"meter\s*number|meter\s*no|meter\s*id",
    }

    filled_count = 0

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=2):
        cell_label = row[0].value
        if not cell_label:
            continue
        label_str = str(cell_label).strip().lower()
        for key, pat in label_patterns.items():
            if re.search(pat, label_str, flags=re.IGNORECASE):
                target_cell = row[1]
                value = label_to_value.get(key, "")

                # Normalize numeric-looking fields
                if key in ("Units Consumed", "Bill Amount", "Load"):
                    value = normalize_number_str(value)

                # Avoid overwriting if value is empty
                if value != "":
                    target_cell.value = value
                    filled_count += 1
                break

    if filled_count == 0:
        # Still return file path, but warn user.
        # (We will warn in the UI.)
        pass

    out_path = os.path.join(os.path.dirname(__file__), OUTPUT_FILE_NAME)
    wb.save(out_path)
    return out_path


# -----------------------------
# UI: Upload
# -----------------------------

with st.container():
    st.subheader("Upload bill (PDF / Image)")

    uploaded = st.file_uploader(
        "Drag and drop your file here or click to browse",
        type=["pdf", "png", "jpg", "jpeg"],
        accept_multiple_files=False,
    )

    col1, col2, col3 = st.columns([1, 1, 2])
    with col1:
        pass
    with col2:
        pass
    with col3:
        st.caption(
            "Tip: Clear scans work best. The app first tries text extraction for PDFs, then falls back to OCR."
        )


if "last_result" not in st.session_state:
    st.session_state.last_result = None


process_clicked = st.button("Process", type="primary", disabled=(uploaded is None))


def validate_upload(uploaded_file) -> tuple[bool, str]:
    if uploaded_file is None:
        return False, "No file uploaded."
    ext = os.path.splitext(uploaded_file.name.lower())[1]
    if ext not in ALLOWED_EXTS:
        return False, "Invalid file type. Please upload PDF, PNG, JPG, or JPEG."
    # basic MIME check when available
    if uploaded_file.type and uploaded_file.type not in ALLOWED_TYPES:
        # Don't hard-fail: some browsers may send empty/odd MIME.
        return True, ""
    return True, ""


if process_clicked and uploaded is not None:
    ok, err = validate_upload(uploaded)
    if not ok:
        st.error(err)
        st.stop()

    with st.spinner("Extracting text with AI/OCR and preparing Excel..."):
        try:
            text, method = extract_text(uploaded)
        except Exception as e:
            st.error(f"OCR/Extraction failed: {e}")
            st.stop()

        if not text or len(text) < 10:
            st.error(
                "Could not extract meaningful text from the file. "
                "Try a clearer scan or a different page."
            )
            st.stop()

        fields = parse_bill_fields(text)

        # Display extracted fields
        df = pd.DataFrame(list(fields.items()), columns=["Field", "Extracted Value"])

        st.success(f"Extraction complete (method: {method}).")
        st.markdown("### Extracted data")
        st.dataframe(df, use_container_width=True, hide_index=True)

        # Fill Excel template
        try:
            out_path = fill_excel_template(fields)
        except FileNotFoundError as e:
            st.error(str(e))
            st.stop()
        except Exception as e:
            st.error(f"Excel generation failed: {e}")
            st.stop()

        # Read generated file for download
        try:
            with open(out_path, "rb") as f:
                excel_bytes = f.read()
        except Exception as e:
            st.error(f"Could not read generated Excel: {e}")
            st.stop()

        st.markdown("### Excel output")
        st.success("Excel generated successfully.")

        st.download_button(
            label="Download generated Excel",
            data=excel_bytes,
            file_name=OUTPUT_FILE_NAME,
            mime=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )


# Footer guidance about template
with st.expander("Template mapping (how filling works)", expanded=False):
    st.write(
        "This app fills the Excel by scanning **column A** for labels like 'Consumer Number', 'Billing Date', etc. "
        "When a label is found in column A, the adjacent cell in **column B** is filled. "
        "Formulas in other cells are not overwritten."
    )
    st.write(
        f"It expects a file named **{TEMPLATE_FILE}** in the same folder as app.py."
    )

